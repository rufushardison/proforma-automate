"""
excel_writer.py

Writes extracted deal assumptions into a copy of the Excel template and
resolves circular references for lender fees and interest carry.

Supports both:
  - Single-tenant templates (flat assumption dict, year-by-year rent in Deal Summary)
  - Multi-tenant templates (rent_roll section in manifest, tenants list from extractor)

Key design decisions:
  - Templates are NEVER modified — always copied to a BytesIO buffer first.
  - keep_vba=True preserves the .xlsm container (VBA macros are stripped by
    openpyxl at runtime but the file remains a valid xlsm).
  - Null/missing assumptions leave the template's default value untouched.
  - Circular-ref cells (lender_fee, interest_carry) are computed by
    circular_solver.py and written after all other assumptions.
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter

from circular_solver import solve_all
from extractor import ExtractionResult

# Yellow  — Claude extracted a value but flagged it as low-confidence
_LOW_CONFIDENCE_FILL = PatternFill(
    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
)

# Orange — Claude found nothing; cell keeps its template default
_NOT_POPULATED_FILL = PatternFill(
    start_color="FFC000", end_color="FFC000", fill_type="solid"
)


# ---------------------------------------------------------------------------
# Value coercion
# ---------------------------------------------------------------------------

def _get_cell_value(raw_value: Any, cell_type: str) -> Any:
    """Coerce a Claude-extracted value to the correct Python type for openpyxl."""
    if raw_value is None:
        return None
    try:
        if cell_type in ("currency", "percent", "float"):
            return float(raw_value)
        elif cell_type == "integer":
            return int(raw_value)
        elif cell_type == "date":
            # Claude returns ISO strings "YYYY-MM-DD"; openpyxl accepts them
            import datetime
            if isinstance(raw_value, str):
                return datetime.datetime.strptime(raw_value[:10], "%Y-%m-%d")
            return raw_value
        else:
            return raw_value  # string — write as-is
    except (TypeError, ValueError):
        return raw_value


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def fill_template(
    template_path: str | Path,
    manifest: dict[str, Any],
    extraction: ExtractionResult,
) -> io.BytesIO:
    """
    Load the template, write assumptions, resolve circular refs, write rent roll
    (for multi-tenant), and return an in-memory .xlsm buffer for download.

    Args:
        template_path: Absolute path to the original .xlsm template.
        manifest:      Parsed JSON manifest dict for this template.
        extraction:    Output of extractor.extract_assumptions().

    Returns:
        BytesIO containing the filled workbook.
    """
    template_path = Path(template_path)
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = openpyxl.load_workbook(template_path, keep_vba=True)

    assumptions_meta = manifest.get("assumptions", {})
    circular_ref_cells = manifest.get("circular_ref_cells", {})

    # ------------------------------------------------------------------
    # 1. Write flat assumptions; highlight every mapped cell
    # ------------------------------------------------------------------
    for assumption_key, cell_meta in assumptions_meta.items():
        extraction_entry = extraction.get(assumption_key, {})
        raw_value = extraction_entry.get("value")
        confidence = extraction_entry.get("confidence", "low")

        sheet_name = cell_meta["sheet"]
        cell_address = cell_meta["cell"]
        cell_type = cell_meta.get("type", "string")

        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]

        if raw_value is None:
            # Nothing extracted — keep template default but mark orange
            ws[cell_address].fill = _NOT_POPULATED_FILL
        else:
            ws[cell_address] = _get_cell_value(raw_value, cell_type)
            if confidence == "low":
                ws[cell_address].fill = _LOW_CONFIDENCE_FILL

    # ------------------------------------------------------------------
    # 2. Write multi-tenant rent roll (if manifest has rent_roll section)
    # ------------------------------------------------------------------
    if "rent_roll" in manifest:
        tenants = extraction.get("_tenants", [])
        if tenants:
            _write_rent_roll(wb, manifest["rent_roll"], tenants)

    # ------------------------------------------------------------------
    # 3. Resolve and write circular reference values
    # ------------------------------------------------------------------
    if circular_ref_cells:
        _write_circular_ref_values(wb, manifest, extraction)

    # ------------------------------------------------------------------
    # 4. Serialize to BytesIO
    # ------------------------------------------------------------------
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------
# Multi-tenant rent roll writer
# ---------------------------------------------------------------------------

def _write_rent_roll(
    wb: openpyxl.Workbook,
    rent_roll_meta: dict[str, Any],
    tenants: list[dict[str, Any]],
) -> None:
    """
    Write per-tenant rows to the Rent Roll sheet.

    Strategy:
      - Write per-tenant fields (name, unit, SF, rent PSF, TI, escalation, NNN flags)
      - Fill all monthly PSF cells for years 2026-2031 with the base_rent_psf
      - Write annual rent dollar amounts to DD-DJ columns (SF × PSF)
    """
    sheet_name = rent_roll_meta.get("sheet", "Rent Roll")
    if sheet_name not in wb.sheetnames:
        return

    ws = wb[sheet_name]
    first_row = rent_roll_meta.get("first_tenant_row", 11)
    last_row = rent_roll_meta.get("last_tenant_row", 27)
    col_map = rent_roll_meta.get("columns", {})
    monthly_cols = rent_roll_meta.get("monthly_psf_columns", {})
    annual_cols = rent_roll_meta.get("annual_rent_columns", {})

    for idx, tenant in enumerate(tenants):
        row = first_row + idx
        if row > last_row:
            break  # no more rows available

        psf = tenant.get("base_rent_psf")
        sf = tenant.get("sq_ft")
        annual_rent = (float(psf) * int(sf)) if psf and sf else None

        # Per-tenant scalar fields
        field_values = {
            "lease_term_years":    (tenant.get("lease_term_years"), "integer"),
            "tenant_name":         (tenant.get("tenant_name"), "string"),
            "unit":                (tenant.get("unit"), "string"),
            "sq_ft":               (tenant.get("sq_ft"), "integer"),
            "base_rent_psf":       (psf, "float"),
            "ti_psf":              (tenant.get("ti_psf"), "float"),
            "annual_pct_increase": (tenant.get("annual_pct_increase"), "float"),
            "tax_reimbursement":   (tenant.get("tax_reimbursement"), "string"),
            "ins_reimbursement":   (tenant.get("ins_reimbursement"), "string"),
            "cam_reimbursement":   (tenant.get("cam_reimbursement"), "string"),
        }

        for field_key, (value, cell_type) in field_values.items():
            col_letter = col_map.get(field_key)
            if not col_letter or value is None:
                continue
            cell = ws[f"{col_letter}{row}"]
            cell.value = _get_cell_value(value, cell_type)
            if tenant.get("confidence") == "low":
                cell.fill = _LOW_CONFIDENCE_FILL

        # Monthly PSF cells — fill every month column with the base_rent_psf
        if psf is not None:
            psf_float = float(psf)
            for year, year_meta in monthly_cols.items():
                start_col = column_index_from_string(year_meta["start"])
                end_col = column_index_from_string(year_meta["end"])
                for col_idx in range(start_col, end_col + 1):
                    ws.cell(row=row, column=col_idx).value = psf_float

        # Annual rent dollar totals (DD-DJ columns)
        if annual_rent is not None:
            for year, col_letter in annual_cols.items():
                ws[f"{col_letter}{row}"].value = annual_rent


# ---------------------------------------------------------------------------
# Circular reference solver
# ---------------------------------------------------------------------------

def _write_circular_ref_values(
    wb: openpyxl.Workbook,
    manifest: dict[str, Any],
    extraction: ExtractionResult,
) -> None:
    """
    Compute lender_fee and interest_carry using circular_solver and write
    to all circular_ref_cells defined in the manifest.

    Uses assumption keys from extraction to gather inputs. Tolerates missing
    values by skipping the solve gracefully.
    """
    def _get(key: str) -> float | None:
        entry = extraction.get(key, {})
        val = entry.get("value")
        if val is None:
            return None
        try:
            return float(val)
        except (TypeError, ValueError):
            return None

    # Gather required inputs — keys are the same across both templates
    acquisition_cost = _get("acquisition_cost")
    ltc = _get("ltc")
    fee_rate = _get("lender_fee_rate")
    rate = _get("financing_rate")
    months_construction = _get("months_of_construction") or 0.0
    dev_fee_rate = _get("developer_fee_rate") or 0.0

    # Also pull all non-circular soft cost line items to compute base_costs
    assumptions_meta = manifest.get("assumptions", {})
    soft_cost_keys = [
        "ae_costs", "contingency", "closing_costs", "permitting", "legal_fees",
        "leasing_commission", "leasing_commission_flat_fee", "additional_cost_amount",
        "land_sale_broker_fee", "tenant_improvement_total",
        "additional_cost_1", "additional_cost_2", "additional_cost_3",
        "additional_cost_4", "additional_cost_5", "additional_cost_6",
        "additional_cost_7", "additional_cost_8", "additional_cost_9",
        "additional_cost_10",
    ]
    soft_costs_total = sum(
        float(extraction[k]["value"])
        for k in soft_cost_keys
        if k in extraction and extraction[k].get("value") is not None
    )

    # For multi-tenant, also add TI costs from Rent Roll tenants
    tenants = extraction.get("_tenants", [])
    for t in tenants:
        ti_psf = t.get("ti_psf") or 0
        sf = t.get("sq_ft") or 0
        try:
            soft_costs_total += float(ti_psf) * int(sf)
        except (TypeError, ValueError):
            pass

    # Also add construction/site-work costs from Costs sheet
    costs_keys = [k for k in assumptions_meta if k.startswith("costs_") or k.startswith("dd_")]
    costs_total = sum(
        float(extraction[k]["value"])
        for k in costs_keys
        if k in extraction and extraction[k].get("value") is not None
    )

    if any(v is None for v in [acquisition_cost, ltc, fee_rate, rate]):
        return  # insufficient inputs — leave template defaults

    base_costs = acquisition_cost + soft_costs_total + costs_total

    try:
        solved = solve_all(
            base_costs=base_costs,
            loan_to_cost=ltc,
            fee_rate=fee_rate,
            annual_rate=rate,
            months_of_construction=months_construction,
            dev_fee_rate=dev_fee_rate,
        )
    except (ValueError, RuntimeError):
        return

    # Write to all circular_ref_cells entries in the manifest
    circular_ref_cells = manifest.get("circular_ref_cells", {})
    _solved_key_map = {
        "lender_fee":     "lender_fee",
        "interest_carry": "interest_carry",
        "loan_amount":    "loan_amount",
    }

    for manifest_key, cell_info in circular_ref_cells.items():
        # manifest_key examples: "lender_fees_deal_summary", "interest_carry_financing"
        solved_key = None
        for k, v in _solved_key_map.items():
            if k in manifest_key:
                solved_key = v
                break
        if solved_key is None:
            continue

        solved_value = solved.get(solved_key)
        if solved_value is None:
            continue

        sheet_name = cell_info.get("sheet", wb.sheetnames[0])
        cell_address = cell_info.get("cell")
        if not cell_address or sheet_name not in wb.sheetnames:
            continue

        wb[sheet_name][cell_address] = solved_value


# ---------------------------------------------------------------------------
# CLI helper
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import json
    import sys

    if len(sys.argv) < 3:
        print("Usage: python excel_writer.py <template.xlsm> <manifest.json>")
        sys.exit(1)

    with open(sys.argv[2]) as f:
        _manifest = json.load(f)

    _extraction: ExtractionResult = {}
    for k, v in _manifest.get("assumptions", {}).items():
        t = v.get("type", "string")
        _extraction[k] = {
            "value": (5_000_000 if t == "currency" else
                      0.065    if t == "percent"  else
                      24       if t == "integer"  else
                      "2026-01-01" if t == "date" else "test"),
            "confidence": "high",
            "note": "",
        }

    buf = fill_template(sys.argv[1], _manifest, _extraction)
    out_path = "output_test.xlsm"
    with open(out_path, "wb") as f:
        f.write(buf.read())
    print(f"Written to {out_path}")

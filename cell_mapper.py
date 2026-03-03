"""
cell_mapper.py

Standalone utility for the cell-mapping session.

Scans a .xlsm (or .xlsx) file for blue-filled cells and prints a JSON
manifest skeleton that you can review, correct, and save as the confirmed
manifest for that template.

Usage:
    python cell_mapper.py path/to/template.xlsm

Optional flags:
    --hex  <RRGGBB>   Exact hex color to match (default: auto-detect top blues)
    --out  <file.json> Write output to file instead of stdout
    --sheet <name>     Scan only a specific sheet

The script will:
  1. Iterate every cell in every sheet (or the named sheet)
  2. Collect cells whose fill foreground color matches the target hex
  3. For each match, look left / above for the nearest non-empty label cell
  4. Print a JSON manifest skeleton with sheet, cell address, and adjacent label
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _normalise_hex(color_obj) -> str | None:
    """
    Extract and normalise a 6-char RRGGBB hex string from an openpyxl Color.
    Returns None if the fill is empty/none.
    """
    if color_obj is None:
        return None
    color_type = getattr(color_obj, "type", None)
    if color_type == "theme":
        return None  # theme colours can't be compared by hex
    rgb = getattr(color_obj, "rgb", None)
    if rgb is None or rgb in ("00000000", "FFFFFFFF", "FF000000"):
        return None
    # openpyxl rgb is AARRGGBB (8 chars) — strip alpha
    if len(rgb) == 8:
        return rgb[2:].upper()
    return rgb.upper()


def _find_label(ws, row: int, col: int) -> str:
    """
    Look left then up from (row, col) to find the nearest non-empty cell value.
    Returns the value as a string, or "" if none found within 3 cells.
    """
    for offset in range(1, 4):
        # Try left
        if col - offset >= 1:
            val = ws.cell(row=row, column=col - offset).value
            if val not in (None, ""):
                return str(val).strip()
    for offset in range(1, 4):
        # Try above
        if row - offset >= 1:
            val = ws.cell(row=row - offset, column=col).value
            if val not in (None, ""):
                return str(val).strip()
    return ""


def collect_colored_cells(
    wb: openpyxl.Workbook,
    target_hex: str | None = None,
    sheet_name: str | None = None,
) -> tuple[list[dict], Counter]:
    """
    Scan workbook for cells with a fill color.

    If target_hex is None, scan all non-white/black fills and return a
    Counter of all found hex values so the user can identify the blue hex.

    If target_hex is given, return only cells matching that hex.

    Returns:
        (matches, all_color_counts)
    """
    matches: list[dict] = []
    all_colors: Counter = Counter()

    sheets = [wb[sheet_name]] if sheet_name else wb.worksheets

    for ws in sheets:
        for row in ws.iter_rows():
            for cell in row:
                fill = cell.fill
                if fill is None or fill.fill_type in (None, "none"):
                    continue
                hex_color = _normalise_hex(fill.fgColor)
                if hex_color is None:
                    continue
                all_colors[hex_color] += 1
                if target_hex and hex_color.upper() != target_hex.upper():
                    continue
                label = _find_label(ws, cell.row, cell.column)
                matches.append(
                    {
                        "sheet": ws.title,
                        "cell": cell.coordinate,
                        "row": cell.row,
                        "col": cell.column,
                        "col_letter": get_column_letter(cell.column),
                        "fill_hex": hex_color,
                        "sample_value": cell.value,
                        "label": label,
                    }
                )

    return matches, all_colors


def build_manifest_skeleton(
    template_file: str,
    matches: list[dict],
) -> dict:
    """
    Build a manifest skeleton from the list of matched cells.
    Assumption keys are auto-generated from the label (snake_case).
    """
    import re

    def to_key(label: str, sheet: str, coord: str) -> str:
        raw = label or f"{sheet}_{coord}"
        key = re.sub(r"[^a-z0-9]+", "_", raw.lower()).strip("_")
        return key or coord.lower()

    assumptions: dict[str, dict] = {}
    seen_keys: dict[str, int] = {}

    for m in matches:
        base_key = to_key(m["label"], m["sheet"], m["cell"])
        count = seen_keys.get(base_key, 0)
        key = base_key if count == 0 else f"{base_key}_{count}"
        seen_keys[base_key] = count + 1

        assumptions[key] = {
            "sheet": m["sheet"],
            "cell": m["cell"],
            "type": "FILL_IN_TYPE",  # currency | percent | integer | float | string
            "required": True,
            "label": m["label"],
            "sample_value": m["sample_value"],
        }

    return {
        "template_name": "FILL_IN_NAME",
        "template_file": Path(template_file).name,
        "circular_ref_cells": {
            "lender_fee":     "CELL_TBD",
            "interest_carry": "CELL_TBD",
            "loan_amount":    "CELL_TBD",
        },
        "_notes": (
            "Review each assumption: correct the key name, set the correct type "
            "(currency/percent/integer/float/string), mark required, "
            "and fill in circular_ref_cells addresses."
        ),
        "assumptions": assumptions,
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Scan a .xlsm file for blue-filled assumption cells and "
                    "output a JSON manifest skeleton."
    )
    parser.add_argument("template", help="Path to the .xlsm (or .xlsx) template file")
    parser.add_argument(
        "--hex",
        metavar="RRGGBB",
        help="Exact 6-char hex of the blue fill color. "
             "Omit to run auto-detection mode.",
    )
    parser.add_argument(
        "--out",
        metavar="FILE",
        help="Write JSON manifest to this file (default: stdout)",
    )
    parser.add_argument(
        "--sheet",
        metavar="NAME",
        help="Scan only this sheet (default: all sheets)",
    )
    args = parser.parse_args()

    tpl_path = Path(args.template)
    if not tpl_path.exists():
        print(f"ERROR: File not found: {tpl_path}", file=sys.stderr)
        sys.exit(1)

    print(f"Loading workbook: {tpl_path} ...", file=sys.stderr)
    wb = openpyxl.load_workbook(tpl_path, keep_vba=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}", file=sys.stderr)

    if not args.hex:
        # Auto-detect: show all colors found and their counts
        _, all_colors = collect_colored_cells(wb, sheet_name=args.sheet)
        if not all_colors:
            print("No colored cells found in this workbook.", file=sys.stderr)
            sys.exit(0)
        print("\nAll fill colors detected (hex: count):", file=sys.stderr)
        for hex_val, count in all_colors.most_common():
            print(f"  {hex_val}  ({count} cells)", file=sys.stderr)
        print(
            "\nRe-run with --hex <RRGGBB> to extract cells of that color.",
            file=sys.stderr,
        )
        sys.exit(0)

    target = args.hex.upper().lstrip("#")
    if len(target) != 6:
        print(f"ERROR: --hex must be exactly 6 hex chars (RRGGBB), got: {target}", file=sys.stderr)
        sys.exit(1)

    matches, _ = collect_colored_cells(wb, target_hex=target, sheet_name=args.sheet)

    if not matches:
        print(
            f"No cells found with fill color #{target}. "
            "Re-run without --hex to see all detected colors.",
            file=sys.stderr,
        )
        sys.exit(0)

    print(f"\nFound {len(matches)} cells with fill #{target}:", file=sys.stderr)
    for m in matches:
        print(
            f"  {m['sheet']!s:<20} {m['cell']:<6}  label={m['label']!r:<30}  "
            f"value={m['sample_value']}",
            file=sys.stderr,
        )

    manifest = build_manifest_skeleton(args.template, matches)
    output_json = json.dumps(manifest, indent=2)

    if args.out:
        out_path = Path(args.out)
        out_path.write_text(output_json)
        print(f"\nManifest written to {out_path}", file=sys.stderr)
    else:
        print("\n" + "=" * 60, file=sys.stderr)
        print("MANIFEST SKELETON (copy to manifests/ and edit):", file=sys.stderr)
        print("=" * 60, file=sys.stderr)
        print(output_json)


if __name__ == "__main__":
    main()
